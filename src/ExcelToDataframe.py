import pandas as pd
from openpyxl import load_workbook


def excel_to_dataframes(file_path):
    """
    Reads all sheets in an Excel file and outputs a dictionary of pandas DataFrames.

    :param file_path: str, path to the Excel file
    :return: dict, a dictionary with sheet names as keys and pandas DataFrame objects as values
    """

    # Load the workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)

    # Initialize a dictionary to store DataFrames
    dfs = {}

    # Iterate through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get the data from the worksheet
        data = []
        for row in ws.iter_rows():
            data.append([cell.value for cell in row])

        # Create a DataFrame from the data
        df = pd.DataFrame(data)

        # Add the DataFrame to the dictionary with the sheet name as the key
        dfs[sheet_name] = df

    # Close the workbook and return the dictionary of DataFrames
    wb.close()
    return dfs
